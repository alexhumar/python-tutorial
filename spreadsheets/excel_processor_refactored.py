import openpyxl as xl
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.chart import BarChart, Reference


def process_workbook(filename: str, extension: str):
    wb = xl.load_workbook(filename + extension)
    sheet: Workbook = wb['Sheet1']

    # Updating prices
    column_number = 3
    first_row = 2
    corrected_column_number = 4
    # range function doesn't include the value of the second argument
    for row in range(first_row, sheet.max_row + 1):
        cell: Cell = sheet.cell(row, column_number)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, corrected_column_number)
        corrected_price_cell.value = corrected_price

    # Creating a bar chart

    # We use the Reference Class to select a range of values
    values = Reference(sheet,
                       min_row=first_row,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename + '_corrected' + extension)


process_workbook(filename='transactions', extension='.xlsx')
